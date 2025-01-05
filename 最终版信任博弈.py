# -*- coding: utf-8 -*-
"""
Created on Wed Dec 23 00:42:45 2024

@author: 18458284448
"""
# 请使用
import matplotlib
matplotlib.use('TkAgg')
import random
import tkinter as tk
from openpyxl import Workbook, load_workbook
import os
import time 

# 加几行空格以使显示得不要太上面
def space(line):
    for _ in range(line):
      print(" ")

# def get_terminal_width():
#     return os.get_terminal_size().columns

# 获取pycharm环境中当前终端的宽度
import os
try:
    from pycharm_display import get_editor_width
    def get_terminal_width():
        return get_editor_width()
except ImportError:
    # 如果没有安装插件或者导入失败，采用前面的容错方式获取宽度
    def get_terminal_width():
        try:
            return os.get_terminal_size().columns
        except OSError:
            return 80

# 居中靠侧显示文本
def print_centered(text):
    terminal_width = get_terminal_width()
    padding = (terminal_width ) // 6 
    print(' ' * padding + text )

    
# 生成包含100个矩形的矩阵，有6种不同颜色
def generate_rectangle_matrix():
    colors = ["red", "blue", "green", "yellow", "purple", "orange"]
    matrix = []
    for _ in range(10):
        row = []
        for _ in range(10):
            color = random.choice(colors)
            row.append(color)
        matrix.append(row)
    return matrix

# 计算矩阵中红色矩形的数量
def count_red_rectangles(matrix):
    count = 0
    for row in matrix:
        for color in row:
            if color == "red":
                count += 1
    return count


# 创建GUI窗口显示矩阵
def display_matrix(matrix):

    root = tk.Tk()
    root.title("矩形判断任务")


    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # 这里假设希望窗口出现在屏幕水平方向居中，垂直方向靠上一些的位置，你可以根据需求调整具体计算的偏移量
    x = (screen_width - len(matrix[0]) * 65) // 2
    y = (screen_height - len(matrix) * 65 - 200) // 2 
    root.geometry("+{}+{}".format(x, y))
    
    for i, row in enumerate(matrix):
        for j, color in enumerate(row):
            canvas = tk.Canvas(root, width=65, height=65, bg=color)
            canvas.grid(row=i, column=j)


    # 在1500ms后关闭窗口
    root.after(1500, root.destroy)

    root.mainloop()

def clear_screen():
        os.system('cls')  # 对于Windows系统



# 显示问题并获取用户输入
def prompt_user(question):
    print_centered("")
    print_centered(question) 
    while True:
        user_input_str = input()
        if user_input_str:  # 判断输入是否为空字符串
            try:
                user_input = int(user_input_str)
                break
            except ValueError:
                print("请输入有效的数字，请重新输入。")
        else:
            print("您不能直接按回车键，请输入相应的数据。")
    clear_screen()
    return user_input

# 进行矩形判断任务
def rectangle_judgment_task():
    total_overconfidence = 0
    for x in range(5):  # 执行5次任务
        matrix = generate_rectangle_matrix()
        display_matrix(matrix)

        # 参与者输入对红色矩形数量的猜测
        space(6)
        print_centered("请输入你对红色矩形数量的猜测(数字): ")
        guess = prompt_user("注:输入数字后单击Enter键即可")

        # 参与者输入对猜测值与实际值差距的估计
        space(8)
        estimated_gap = prompt_user("请输入你认为你的猜测数量与正确答案红色矩形数量的差距: ")
        if x!= 4:
               space(7)
               print_centered("请按回车键开始下一矩阵的测试，按下1s后矩阵出现")
               input("")
               clear_screen()
               time.sleep(1)
        # 计算实际红色矩形数量
        actual_count = count_red_rectangles(matrix)

        # 计算实际差距
        actual_gap = abs(actual_count - guess)

        # 计算过度自信程度（实际差距 - 估计差距）
        overconfidence = actual_gap - estimated_gap
        total_overconfidence += overconfidence
    average_overconfidence = total_overconfidence / 5
    return average_overconfidence

# 信任游戏函数
def trust_game():
    # 模拟合作方（虚拟玩家）的策略，以牙还牙策略
    def partner_strategy(previous_investment, current_investment, base_return_percentage=0.5):
        if current_investment == 0 and previous_investment == 0:
            return_percentage = 0.5  # 第一轮且双方投资都为0时，使用基础回馈比例
        elif current_investment == 0:
            return_percentage = max(0.25, base_return_percentage - 0.2)  # 如果当前投资额为0（非第一轮），回馈比例大幅降低，但不低于0.25
        elif current_investment > previous_investment:
            return_percentage = min(0.75, base_return_percentage + 0.1)  # 如果投资额增加，回馈比例增加，但不高于0.75
        elif current_investment < previous_investment:
            return_percentage = max(0.25, base_return_percentage - 0.1)  # 如果投资额减少，回馈比例降低，但不低于0.25
        else:
            return_percentage = base_return_percentage
        return return_percentage

    # 记录投资方每轮投资额和最终收益
    investments = []
    total_earnings = 0
    previous_investment = 0  # 初始化上一轮投资额为0，第一轮视为上一轮投资额为0的情况
    base_return_percentage = 0.5  # 明确基础回馈比例变量
    
    for round in range(10):
        initial_fund = 100
        if round == 0:
            space(8)
            print_centered(f" 现在任务正式开始，现在是第{round + 1}轮：")
            investment = prompt_user(f"你本轮启动基金为{initial_fund}元，你愿意拿出几元向对方投资? ")
        else:
            investment = prompt_user(f" 第{round + 1}轮：你本轮启动资金依然为{initial_fund}元，你愿意拿出几元向对方投资？ ")
        # 确保投资金额不超过当前拥有资金
        investment = min(investment, 100)
        investments.append(investment)
        initial_fund -= investment

        # 合作方获得收益并计算回馈
        partner_earnings = investment * 3
        space(6)
        print_centered(f"对方获得您投资的三倍，即{partner_earnings}元，对方正在决定返还的钱数，请稍后...")
        
        random_seconds = random.randint(2, 5)
        time.sleep(random_seconds)
        
        if round == 0:
            partner_return_percentage = base_return_percentage  # 第一轮使用基础回馈比例
        else:
            partner_return_percentage = partner_strategy(previous_investment, investment, base_return_percentage)
            base_return_percentage = partner_return_percentage  # 更新基础回馈比例，为下一轮做准备
        partner_return = int(partner_earnings * partner_return_percentage)


        initial_fund += partner_return
        total_earnings += initial_fund
        
        space(2)
        print_centered(f"对方决定从{partner_earnings}元收益中给你{partner_return}元。")
        print_centered(f"    你在本轮共获得100 - {investment} + {partner_return} = {initial_fund}元。")
        space(2)
        print_centered(f"您已共计进行{round + 1}轮博弈，累计收益为{total_earnings}元")
        print_centered("请注意，您下轮的起始资金依然为100元，每轮初始资金是独立的")
        space(1)
        print_centered("请按Enter键继续...")
        input("")
        clear_screen() 
        space(3)
        # 更新上一轮投资额，为下一轮计算做准备
        previous_investment = investment

    print_centered(f" 游戏结束，你的总收益为{total_earnings}元。")
    space(2)
    print_centered(" 实验结束，感谢您的参与！请按Enter键上传数据")
    input("")
    return investments

if __name__ == "__main__":
    # 检查是否存在已有的Excel文件，若不存在则创建新的工作簿和工作表
    try:
        workbook = load_workbook('experiment_data.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['过度自信程度', '自我效能感', '投资金额1', '投资金额2', '投资金额3', '投资金额4', '投资金额5', '投资金额6', '投资金额7', '投资金额8', '投资金额9', '投资金额10'])

    space(8)
    print_centered("很感谢您来参与本次实验，请认真阅读指导语。")
    print_centered("    这是一个合作博弈投资任务，我们将首先通过一个估计红色矩形数量任务测量您的计算能力，")
    print_centered("接下来请您按Enter键进入实验，按完后会立即呈现由10*10个不同颜色方块组成的一个矩阵")
    print_centered("    请在其消失前尽可能估计其中红色方块数量")
    space(1)
    print_centered("    确认明白任务要求后，请按Enter键继续，紧接着会立即展示颜色矩阵，请注意估算")
    print_centered("    按下按键后，矩阵将在一秒后出现在屏幕中央，持续1.5秒")
    
    input("")
    clear_screen()
    time.sleep(1)
  
    average_overconfidence = rectangle_judgment_task()
    clear_screen()
    space(8)
    print_centered("    接下来，您将有机会在一个投资博弈任务中大显身手，这是一次真实的投资任务。")
    print_centered("游戏将在投资方(A方)和接受资金方(B方)之间展开。您将随机分配到一方")
    space(1)
    print_centered("按Enter继续...")
    input("")
    clear_screen()
    space(10)
    print_centered("分配中...    请勿操作")
    time.sleep(4)
    clear_screen()
    space(8)
    print_centered("    您的角色为投资方(A)，规则如下：实验中，您在每轮投资初始拥有100元人民币，")
    print_centered("您需要选择X元交给对方角色(B)进行打理，他同你一样为真实参与者") 
    print_centered("    您的投资会获得3X元的收益，对方将选择一部分或全部返还给您，给您Y元，当然对方也") 
    print_centered("可以选择独吞，因此请慎重思考")
    space(1)
    print_centered("也就是说在一轮游戏中，您的收益为：100-X+Y元。单独每轮的收益独立。") 
    print_centered("    请您认真思考后再给予投资，游戏共进行10轮，每轮相加的最终累计获得金额将与您获") 
    print_centered("得的被试费相挂钩") 
    space(1)
    print_centered("    请按Enter键继续...")
    input("")
    # 询问被试对完成信任游戏任务的信心程度
    space(5)
    print_centered("  在任务开始前，请您回答以下一个问题：")
    confidence = prompt_user("我认为自己有能力很好地完成这次本次投资博弈任务（1完全不正确 - 7分完全正确）？")
    clear_screen()
    investments = trust_game()
   
    # 将数据写入Excel表格
    data = [average_overconfidence, confidence] + investments
    sheet.append(data)
    

    workbook.save('experiment_data.xlsx')
 